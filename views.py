from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.shortcuts import redirect
from django.http import HttpResponse, JsonResponse
from datetime import datetime, date, timedelta
from schedules.models import Schedule, Subject
from groups_app.models import Student, Group
from attendance.models import Attendance, ExcuseRequest, QRSession
from users.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
import io
import json
from django.views.decorators.csrf import csrf_exempt
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
import io
from datetime import datetime


def get_weekday_name(weekday):
    days = {
        1: 'Понедельник',
        2: 'Вторник',
        3: 'Среда',
        4: 'Четверг',
        5: 'Пятница',
        6: 'Суббота',
        7: 'Воскресенье',
    }
    return days.get(weekday, '')


@login_required
def dashboard(request):
    user = request.user
    role = user.role
    context = {'role': role}

    if role == 'teacher':
        today = date.today()
        weekday = today.isoweekday()
        lessons = Schedule.objects.filter(teacher=user, weekday=weekday)
        context['lessons'] = lessons
        context['today'] = today
        context['weekday_name'] = get_weekday_name(weekday)
        print(f"Учитель {user.username}, пар сегодня: {lessons.count()}")  # Отладка

    elif role == 'student':
        try:
            student = Student.objects.get(user=user)
            lessons = Schedule.objects.filter(group=student.group)
            attendances = Attendance.objects.filter(student=student)
            total = attendances.count()
            present = attendances.filter(status='present').count()
            percent = (present / total * 100) if total > 0 else 0
            context['lessons'] = lessons
            context['student'] = student
            context['percent'] = round(percent)
        except Student.DoesNotExist:
            context['lessons'] = []
            context['percent'] = 0

    elif role == 'head':
        groups = Group.objects.all()
        lessons = Schedule.objects.all()
        context['groups'] = groups
        context['lessons'] = lessons

    elif role == 'admin':
        groups = Group.objects.all()
        lessons = Schedule.objects.all()
        context['groups'] = groups
        context['lessons'] = lessons

    else:
        lessons = Schedule.objects.all()
        context['lessons'] = lessons

    print(f"Контекст: role={role}, lessons={context.get('lessons', [])}")  # Отладка
    return render(request, 'dashboard.html', context)


@login_required
def schedule_page(request):
    user = request.user
    role = user.role
    context = {'role': role}

    if role == 'teacher':
        schedule_by_day = {}
        for day in range(1, 7):
            lessons = Schedule.objects.filter(teacher=user, weekday=day).order_by('start_time')
            if lessons:
                schedule_by_day[day] = {
                    'name': get_weekday_name(day),
                    'lessons': lessons
                }
        context['schedule_by_day'] = schedule_by_day

    elif role == 'student':
        try:
            student = Student.objects.get(user=user)
            schedule_by_day = {}
            for day in range(1, 7):
                lessons = Schedule.objects.filter(group=student.group, weekday=day).order_by('start_time')
                if lessons:
                    schedule_by_day[day] = {
                        'name': get_weekday_name(day),
                        'lessons': lessons
                    }
            context['schedule_by_day'] = schedule_by_day
        except Student.DoesNotExist:
            context['schedule_by_day'] = {}
    else:
        groups = Group.objects.all()
        context['groups'] = groups

    return render(request, 'schedule_page.html', context)


@login_required
def attendance_page(request):
    user = request.user
    role = user.role
    context = {'role': role}

    if role == 'teacher':
        groups = Group.objects.filter(curator=user)
        context['groups'] = groups

    elif role == 'student':
        try:
            student = Student.objects.get(user=user)
            attendances = Attendance.objects.filter(student=student).order_by('-date')
            total = attendances.count()
            present = attendances.filter(status='present').count()
            late_count = attendances.filter(status='late').count()
            absent_count = attendances.filter(status='absent').count()
            excused_count = attendances.filter(status='excused').count()
            percent = (present / total * 100) if total > 0 else 0
            context['attendances'] = attendances
            context['percent'] = round(percent)
            context['present'] = present
            context['late'] = late_count
            context['absent'] = absent_count
            context['excused'] = excused_count
        except Student.DoesNotExist:
            context['attendances'] = []
            context['percent'] = 0
            context['present'] = 0
            context['late'] = 0
            context['absent'] = 0
            context['excused'] = 0
    else:
        groups = Group.objects.all()
        context['groups'] = groups
        context['now'] = date.today()

    return render(request, 'attendance_page.html', context)


@login_required
def attendance_journal(request):
    user = request.user
    role = user.role
    context = {'role': role}

    if role == 'teacher':
        groups = Group.objects.filter(curator=user)
        context['groups'] = groups

        group_id = request.GET.get('group_id')
        selected_date = request.GET.get('date')
        schedule_id = request.GET.get('schedule_id')

        if group_id and selected_date and schedule_id:
            group = get_object_or_404(Group, id=group_id)
            students = Student.objects.filter(group=group, is_expelled=False)
            schedule = get_object_or_404(Schedule, id=schedule_id)

            context['selected_group'] = group
            context['selected_date'] = selected_date
            context['students'] = students
            context['schedule'] = schedule

    return render(request, 'attendance_journal.html', context)


@login_required
def save_attendance(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_id = data.get('student_id')
            date_str = data.get('date')
            status = data.get('status')
            late_minutes = data.get('late_minutes', 0)
            reason = data.get('reason', '')
            schedule_id = data.get('schedule_id')

            student = get_object_or_404(Student, id=student_id)
            schedule = get_object_or_404(Schedule, id=schedule_id)
            att_date = datetime.strptime(date_str, '%Y-%m-%d').date()

            attendance, created = Attendance.objects.update_or_create(
                student=student,
                schedule=schedule,
                date=att_date,
                defaults={
                    'status': status,
                    'late_minutes': late_minutes if status == 'late' else 0,
                    'reason': reason if status == 'absent' or status == 'excused' else ''
                }
            )

            return JsonResponse({'success': True, 'message': 'Сохранено'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Метод не разрешен'})


@login_required
def mark_all_present(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            group_id = data.get('group_id')
            date_str = data.get('date')
            schedule_id = data.get('schedule_id')

            students = Student.objects.filter(group_id=group_id, is_expelled=False)
            schedule = get_object_or_404(Schedule, id=schedule_id)
            att_date = datetime.strptime(date_str, '%Y-%m-%d').date()

            for student in students:
                Attendance.objects.update_or_create(
                    student=student,
                    schedule=schedule,
                    date=att_date,
                    defaults={'status': 'present', 'late_minutes': 0, 'reason': ''}
                )

            return JsonResponse({'success': True, 'message': f'Отмечено {students.count()} студентов'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Метод не разрешен'})


@login_required
def reports_page(request):
    user = request.user
    role = user.role
    context = {'role': role}

    if role == 'student':
        try:
            student = Student.objects.get(user=user)
            attendances = Attendance.objects.filter(student=student)
            total = attendances.count()
            present = attendances.filter(status='present').count()
            late = attendances.filter(status='late').count()
            excused = attendances.filter(status='excused').count()
            absent = attendances.filter(status='absent').count()
            percent = (present / total * 100) if total > 0 else 0

            context['student'] = student
            context['percent'] = round(percent)
            context['total'] = total
            context['present'] = present
            context['late'] = late
            context['excused'] = excused
            context['absent'] = absent
            context['report_type'] = 'student'
        except Student.DoesNotExist:
            context['report_type'] = 'error'

    elif role == 'teacher':
        groups = Group.objects.filter(curator=user)
        my_lessons = Schedule.objects.filter(teacher=user)
        context['groups'] = groups
        context['my_lessons'] = my_lessons
        context['report_type'] = 'teacher'

    elif role == 'head':
        groups = Group.objects.all()
        context['groups'] = groups
        context['report_type'] = 'head'

    elif role == 'admin':
        groups = Group.objects.all()
        users = User.objects.all()
        context['groups'] = groups
        context['users'] = users
        context['report_type'] = 'admin'

    return render(request, 'reports_page.html', context)


@login_required
def notifications_page(request):
    user = request.user
    role = user.role
    context = {'role': role}
    context['message'] = 'Страница уведомлений'
    return render(request, 'notifications_page.html', context)


from django.views.decorators.csrf import csrf_exempt


@csrf_exempt
def mark_attendance_by_qr(request, schedule_id):
    from datetime import date
    schedule = get_object_or_404(Schedule, id=schedule_id)
    today = date.today()

    if request.method == 'GET':
        return render(request, 'qr_mark.html', {'schedule': schedule})

    if request.method == 'POST':
        if not request.user.is_authenticated:
            return JsonResponse({'success': False, 'message': 'Необходимо войти в систему'})

        if request.user.role != 'student':
            return JsonResponse({'success': False, 'message': 'Только студенты могут отмечаться'})

        try:
            student = Student.objects.get(user=request.user)
        except Student.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Студент не найден'})

        attendance, created = Attendance.objects.get_or_create(
            student=student,
            schedule=schedule,
            date=today,
            defaults={'status': 'present', 'late_minutes': 0}
        )

        if created:
            return JsonResponse({'success': True, 'message': 'Вы успешно отмечены'})
        else:
            return JsonResponse({'success': False, 'message': 'Вы уже отмечены на этой паре'})

    return JsonResponse({'success': False, 'message': 'Метод не разрешен'})


def attendance_stats_api(request):
    group_id = request.GET.get('group_id')
    selected_date = request.GET.get('date')

    if not group_id:
        return JsonResponse({'success': False, 'message': 'Группа не указана'})

    try:
        group = Group.objects.get(id=group_id)
    except Group.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Группа не найдена'})

    students = Student.objects.filter(group=group, is_expelled=False)

    # Если дата не указана, берем сегодня
    if selected_date:
        target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    else:
        target_date = date.today()

    present = 0
    late = 0
    absent = 0
    excused = 0

    for student in students:
        attendance = Attendance.objects.filter(
            student=student,
            date=target_date
        ).first()

        if attendance:
            if attendance.status == 'present':
                present += 1
            elif attendance.status == 'late':
                late += 1
            elif attendance.status == 'absent':
                absent += 1
            elif attendance.status == 'excused':
                excused += 1

    return JsonResponse({
        'success': True,
        'present': present,
        'late': late,
        'absent': absent,
        'excused': excused
    })


@login_required
def create_excuse_request(request):
    if request.method == 'POST':
        try:
            attendance_id = request.POST.get('attendance_id')
            comment = request.POST.get('comment')
            document = request.FILES.get('document')

            attendance = get_object_or_404(Attendance, id=attendance_id)

            if request.user.role == 'student':
                student = Student.objects.get(user=request.user)
                if attendance.student != student:
                    return JsonResponse({'success': False, 'message': 'Доступ запрещен'})

            excuse_request = ExcuseRequest.objects.create(
                attendance=attendance,
                comment=comment,
                document=document,
                approved=None
            )

            return JsonResponse({'success': True, 'message': 'Заявка отправлена'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Метод не разрешен'})


@login_required
def process_excuse_request(request, request_id, action):
    if request.method == 'POST':
        try:
            excuse_request = get_object_or_404(ExcuseRequest, id=request_id)
            teacher_comment = request.POST.get('comment', '')

            if request.user.role not in ['teacher', 'head', 'admin']:
                return JsonResponse({'success': False, 'message': 'Доступ запрещен'})

            if action == 'approve':
                excuse_request.approved = True
                excuse_request.attendance.status = 'excused'
                excuse_request.attendance.reason = excuse_request.comment
                excuse_request.attendance.save()
                message = 'Заявка одобрена'
            elif action == 'reject':
                excuse_request.approved = False
                message = 'Заявка отклонена'
            else:
                return JsonResponse({'success': False, 'message': 'Неверное действие'})

            excuse_request.teacher_comment = teacher_comment
            excuse_request.save()

            return JsonResponse({'success': True, 'message': message})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Метод не разрешен'})


def custom_logout(request):
    logout(request)
    return redirect('/accounts/login/')


# ==================== ОТЧЕТЫ ====================

def export_student_report_excel(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    attendances = Attendance.objects.filter(student=student).order_by('-date')

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Посещаемость студента"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0b4f6c", end_color="0b4f6c", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells('A1:F1')
    ws['A1'] = f"ХПК: Посещаемость - Справка студента"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws['A3'] = f"ФИО студента:"
    ws['B3'] = f"{student.user.get_full_name()}"
    ws['A4'] = f"Группа:"
    ws['B4'] = f"{student.group.name}"
    ws['A5'] = f"Дата поступления:"
    ws['B5'] = f"{student.admission_date}"
    ws['A6'] = f"Дата выдачи справки:"
    ws['B6'] = f"{datetime.now().strftime('%d.%m.%Y')}"

    headers = ['Дата', 'Предмет', 'Преподаватель', 'Статус', 'Опоздание (мин)', 'Причина']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row = 9
    total_present = 0
    for att in attendances:
        ws.cell(row=row, column=1, value=att.date.strftime('%d.%m.%Y')).border = border
        ws.cell(row=row, column=2, value=att.schedule.subject.title).border = border
        ws.cell(row=row, column=3, value=att.schedule.teacher.get_full_name()).border = border
        status_display = dict(Attendance.STATUS_CHOICES).get(att.status, att.status)
        ws.cell(row=row, column=4, value=status_display).border = border
        ws.cell(row=row, column=5, value=att.late_minutes if att.late_minutes else '-').border = border
        ws.cell(row=row, column=6, value=att.reason if att.reason else '-').border = border
        if att.status == 'present':
            total_present += 1
        row += 1

    total = attendances.count()
    percent = round((total_present / total * 100) if total > 0 else 0)
    ws.cell(row=row + 1, column=1, value="ИТОГО:")
    ws.cell(row=row + 1, column=4, value=f"Посещаемость: {percent}%")

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename="spravka_{student.user.last_name}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    workbook.save(response)
    return response


def export_group_report_excel(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    students = Student.objects.filter(group=group, is_expelled=False)

    workbook = Workbook()
    ws = workbook.active
    ws.title = f"Отчет группы {group.name}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0b4f6c", end_color="0b4f6c", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells('A1:G1')
    ws['A1'] = f"ХПК: Посещаемость - Отчет по группе {group.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws['A3'] = f"Группа:"
    ws['B3'] = f"{group.name}"
    ws['A4'] = f"Курс:"
    ws['B4'] = f"{group.course}"
    ws['A5'] = f"Куратор:"
    ws['B5'] = f"{group.curator.get_full_name() if group.curator else 'Не назначен'}"
    ws['A6'] = f"Дата формирования:"
    ws['B6'] = f"{datetime.now().strftime('%d.%m.%Y')}"

    headers = ['№', 'ФИО студента', 'Всего пар', 'Присутствовал', 'Опоздал', 'Уваж. причина', 'Процент']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row = 9
    for i, student in enumerate(students, 1):
        attendances = Attendance.objects.filter(student=student)
        total = attendances.count()
        present = attendances.filter(status='present').count()
        late = attendances.filter(status='late').count()
        excused = attendances.filter(status='excused').count()
        percent = round((present / total * 100) if total > 0 else 0)

        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=student.user.get_full_name()).border = border
        ws.cell(row=row, column=3, value=total).border = border
        ws.cell(row=row, column=4, value=present).border = border
        ws.cell(row=row, column=5, value=late).border = border
        ws.cell(row=row, column=6, value=excused).border = border
        ws.cell(row=row, column=7, value=f"{percent}%").border = border
        row += 1

    ws.cell(row=row + 1, column=1, value="ВСЕГО СТУДЕНТОВ:")
    ws.cell(row=row + 1, column=2, value=students.count())

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename="report_group_{group.name}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    workbook.save(response)
    return response


def export_teacher_report_excel(request):
    user = request.user
    lessons = Schedule.objects.filter(teacher=user)

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Отчет преподавателя"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0b4f6c", end_color="0b4f6c", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells('A1:F1')
    ws['A1'] = f"ХПК: Посещаемость - Отчет преподавателя"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws['A3'] = f"Преподаватель:"
    ws['B3'] = f"{user.get_full_name()}"
    ws['A4'] = f"Дата формирования:"
    ws['B4'] = f"{datetime.now().strftime('%d.%m.%Y')}"

    headers = ['№', 'Предмет', 'Группа', 'Даты пар', 'Всего студентов', 'Средняя посещаемость']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row = 7
    for i, lesson in enumerate(lessons, 1):
        attendances = Attendance.objects.filter(schedule=lesson)
        students = Student.objects.filter(group=lesson.group)
        total_students = students.count()
        total_marks = attendances.count()
        present_marks = attendances.filter(status='present').count()
        avg_percent = round((present_marks / total_marks * 100) if total_marks > 0 else 0)

        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=lesson.subject.title).border = border
        ws.cell(row=row, column=3, value=lesson.group.name).border = border
        ws.cell(row=row, column=4, value=f"{lesson.start_time}-{lesson.end_time}").border = border
        ws.cell(row=row, column=5, value=total_students).border = border
        ws.cell(row=row, column=6, value=f"{avg_percent}%").border = border
        row += 1

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename="report_teacher_{user.last_name}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    workbook.save(response)
    return response


def export_summary_report_excel(request):
    groups = Group.objects.all()

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Сводный отчет по колледжу"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0b4f6c", end_color="0b4f6c", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells('A1:G1')
    ws['A1'] = f"ХПК: Посещаемость - Сводный отчет по колледжу"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws['A3'] = f"Дата формирования:"
    ws['B3'] = f"{datetime.now().strftime('%d.%m.%Y')}"
    ws['A4'] = f"Учебный год:"
    ws['B4'] = f"2025/2026"

    headers = ['№', 'Группа', 'Курс', 'Куратор', 'Студентов', 'Всего отметок', 'Средняя посещаемость']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row = 7
    for i, group in enumerate(groups, 1):
        students = Student.objects.filter(group=group, is_expelled=False)
        attendances = Attendance.objects.filter(student__group=group)
        total_marks = attendances.count()
        present_marks = attendances.filter(status='present').count()
        avg_percent = round((present_marks / total_marks * 100) if total_marks > 0 else 0)

        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=2, value=group.name).border = border
        ws.cell(row=row, column=3, value=group.course).border = border
        ws.cell(row=row, column=4,
                value=group.curator.get_full_name() if group.curator else 'Не назначен').border = border
        ws.cell(row=row, column=5, value=students.count()).border = border
        ws.cell(row=row, column=6, value=total_marks).border = border
        ws.cell(row=row, column=7, value=f"{avg_percent}%").border = border
        row += 1

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="summary_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    workbook.save(response)
    return response


@login_required
def attendance_journal(request):
    user = request.user
    role = user.role
    context = {'role': role}

    if role == 'teacher':
        groups = Group.objects.filter(curator=user)
        context['groups'] = groups

        group_id = request.GET.get('group_id')
        selected_date = request.GET.get('date')
        schedule_id = request.GET.get('schedule_id')

        if group_id and selected_date and schedule_id:
            group = get_object_or_404(Group, id=group_id)
            students = Student.objects.filter(group=group, is_expelled=False)
            schedule = get_object_or_404(Schedule, id=schedule_id)
            att_date = datetime.strptime(selected_date, '%Y-%m-%d').date()

            attendances = {}
            for student in students:
                # Автоматически создаем запись, если её нет (со статусом present)
                att, created = Attendance.objects.get_or_create(
                    student=student,
                    schedule=schedule,
                    date=att_date,
                    defaults={'status': 'present', 'late_minutes': 0, 'reason': ''}
                )
                attendances[student.id] = att

            context['selected_group'] = group
            context['selected_date'] = selected_date
            context['students'] = students
            context['schedule'] = schedule
            context['attendances'] = attendances

    return render(request, 'attendance_journal.html', context)